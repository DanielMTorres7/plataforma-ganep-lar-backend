import io
import pandas as pd

from datetime import datetime
from typing import List, Optional
from cachetools import TTLCache, cached
from flask import jsonify, send_file, Request, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment
from services.mongo import db

# Configuração do cache
cache_a = TTLCache(maxsize=100, ttl=3600)
@cached(cache_a)
def get_atendimentos() -> pd.DataFrame:
    """Obtém os dados de atendimentos_completo e retorna um DataFrame."""
    colecao_atendimentos = db["prontuarios"]
    # Consulta todos os documentos na coleção
    dados = list(colecao_atendimentos.find())
    
    return dados

def get_df(data_inicio: datetime, data_fim: datetime, operadoras: Optional[List[str]] = None) -> tuple[pd.DataFrame, int, pd.DataFrame, int, List[dict]]:
    """Calcula o número de pacientes com ScoreBraden por mês."""
    prontuarios = get_atendimentos()

    lpps = []
    scorebradens = []
    for prontuario in prontuarios:
        for _, atendimento in prontuario['ATENDIMENTOS'].items():
            if pd.isna(atendimento['ENTRADA']) or not atendimento['ENTRADA'] or atendimento['ENTRADA'] == "":
                continue
            if atendimento['STATUS'] not in ['Alta', 'Em atendimento']:
                continue
            if atendimento['ENTRADA'] > (data_fim) or (atendimento['STATUS'] == 'Alta' and atendimento['ALTA'] < data_inicio):
                continue
            # if str(prontuario['PRONTUARIO']) != str(2280):
            #     continue
            # print("asdasd",atendimento['ENTRADA'], atendimento['ALTA'])

            if atendimento['SCORE_BRADEN']:
                scorebradens.append({
                    'PACIENTE': prontuario['PACIENTE'],
                    'PRONTUARIO': prontuario['PRONTUARIO'],
                    'ATENDIMENTO': atendimento['ATENDIMENTO'],
                    'ENTRADA': atendimento['ENTRADA'],
                    'STATUS': atendimento['STATUS'],
                    'ALTA': atendimento['ALTA'],
                    'OPERADORA': atendimento['OPERADORA']
                })

            for _, inter in atendimento['INTERCORRENCIAS'].items():
                if pd.isna(inter['DATA_INICIO']):
                    continue
                if inter['CLASSIFICACAO'] == 'LPP' and inter['DATA_INICIO'] >= data_inicio and inter['DATA_INICIO'] <= data_fim:
                    lpps.append({
                        'PACIENTE': prontuario['PACIENTE'],
                        'PRONTUARIO': prontuario['PRONTUARIO'],
                        'ATENDIMENTO': atendimento['ATENDIMENTO'],
                        'DATA_INICIO': inter['DATA_INICIO'],
                        'OPERADORA': atendimento['OPERADORA']
                    })

    lpp_table = pd.DataFrame(lpps).sort_values(by='DATA_INICIO') if lpps else pd.DataFrame()
    list_operadoras = [{'label': f'{op} : {lpp_table[lpp_table["OPERADORA"] == op].shape[0]}', 'value': op} for op in lpp_table['OPERADORA'].unique()] if not lpp_table.empty else []
    scorebradens = pd.DataFrame(scorebradens)
    if operadoras:
        scorebradens = scorebradens[scorebradens['OPERADORA'].isin(operadoras)]
        lpp_table = lpp_table[lpp_table['OPERADORA'].isin(operadoras)] if not lpp_table.empty else lpp_table
    
    n_lpps = lpp_table.shape[0]  # Número total de LPPs no período
    n_atendimentos = scorebradens.shape[0]

    # Contar mensalmente os pacientes com ScoreBraden no período de interesse
    score_mensal = []
    for mes in pd.date_range(data_inicio, data_fim, freq='MS'):
        inicio = mes
        fim = (mes + pd.DateOffset(months=1))
        # print(inicio, fim, scorebradens['ENTRADA'])
        scorebradens_mes = len(scorebradens[
            (scorebradens['ENTRADA'] < fim) &
            ((scorebradens['STATUS'].isin(['Alta', 'Em atendimento']))) &
            ((pd.isna(scorebradens['ALTA'])) | (scorebradens['ALTA'] >= inicio)) 
        ]['PRONTUARIO'].drop_duplicates())
        
        # pd.DataFrame(scorebradens_mes).to_csv('scorebradens_' + mes.strftime('%m') + '.csv')
        # scorebradens_mes = scorebradens_mes.shape[0]
        lpps_mes = 0 if lpp_table.empty else lpp_table[lpp_table['DATA_INICIO'].dt.to_period('M') == mes.to_period('M')].shape[0]
        percentual = round(lpps_mes / scorebradens_mes * 100, 2) if lpps_mes > 0 else 0
        score_mensal.append(
            {
                'mes': mes,
                'score_braden': scorebradens_mes,
                'lpps': lpps_mes,
                'percentual': percentual
            }
        )

    score_mensal = pd.DataFrame(score_mensal).sort_values(by='mes')
    score_mensal['mes'] = score_mensal['mes'].dt.strftime('%m/%Y')
    return score_mensal, lpp_table, n_lpps, n_atendimentos, list_operadoras

def get_data(request: Request) -> Response:
    """Endpoint para obter dados de LPPs e ScoreBraden."""
    try:
        # Obtém os dados da requisição
        data = request.json

        # Verifica se os dados foram fornecidos
        if not data:
            return jsonify({"error": "Dados não fornecidos no corpo da requisição"}), 400

        # Valida os campos obrigatórios
        inicio = data.get("data_inicio")
        fim = data.get("data_fim")
        operadoras = data.get("operadoras")

        if not inicio or not fim:
            return jsonify({"error": "Os atributos 'data_inicio' e 'data_fim' são obrigatórios"}), 400

        # Converte as datas para o formato datetime
        data_inicio = pd.to_datetime(inicio, format='%Y-%m-%d', errors='coerce')
        data_fim = pd.to_datetime(fim, format='%Y-%m-%d', errors='coerce').replace(hour=23, minute=59, second=59)
        print(data_inicio, data_fim)

        # Verifica se as datas foram convertidas corretamente
        if pd.isna(data_inicio) or pd.isna(data_fim):
            return jsonify({"error": "Formato de data inválido. Use o formato 'YYYY-MM-DD'"}), 400

        # Verifica se a data de início é anterior à data de fim
        if data_inicio > data_fim:
            return jsonify({"error": "A data de início deve ser anterior à data de fim"}), 400

        # Obtém os dados
        score_braden_mensal, lpp_table, n_lpps, score_braden_total, list_operadoras = get_df(data_inicio, data_fim, operadoras)

        # Resposta final
        response_data = {
            'lpp_table': lpp_table.to_dict(orient='records'),
            'operadoras': list_operadoras,
            'score_braden_mensal': score_braden_mensal.to_dict(orient='records'),
            'pacientes_classificados_score_braden': score_braden_total,
            'numero_lpps': n_lpps,  # Número total de LPPs no período
        }

        return jsonify(response_data), 200

    except Exception as e:
        # Log de erro para depuração
        print(f"Erro em get_data: {str(e)}")
        return jsonify({"error": "Erro interno no servidor"}), 500
    
def getatends(data_inicio: datetime, data_fim: datetime, operadoras: Optional[List[str]] = None) -> List[dict]:
    prontuarios = get_atendimentos()
    atendimentos = []
    for prontuario in prontuarios.to_dict(orient='records'):
        for n_atendimento, atendimento in prontuario['ATENDIMENTOS'].items():
            if atendimento['STATUS'] not in ['Alta', 'Em atendimento']:
                continue
            if atendimento['ENTRADA'] > data_fim or (atendimento['STATUS'] == 'Alta' and atendimento['ALTA'] < data_inicio):
                continue

            for _, inter in atendimento['INTERCORRENCIAS'].items():
                if inter and inter['CLASSIFICACAO'] == 'LPP' and data_inicio <= inter['DATA_INICIO'] <= data_fim:
                    atendimentos.append({
                        'PACIENTE': prontuario['PACIENTE'],
                        'PRONTUARIO': prontuario['PRONTUARIO'],
                        'ATENDIMENTO': n_atendimento,
                        'DATA_INICIO': inter['DATA_INICIO'],
                        'OPERADORA': atendimento['OPERADORA'],
                        'STATUS': atendimento['STATUS'],
                        'DATA_ALTA': atendimento['ALTA']
                    })
    if operadoras:
        atendimentos = [atend for atend in atendimentos if atend['OPERADORA'] in operadoras]

    return sorted(atendimentos, key=lambda x: x['DATA_INICIO'])

def download_xlsx(request: Request) -> Response:
    data = request.json

    if not data:
        return jsonify({"error": "Dados não fornecidos no corpo da requisição"}), 400

    inicio = data.get("data_inicio")
    fim = data.get("data_fim")
    operadoras = data.get("operadoras")

    if not inicio or not fim:
        return jsonify({"error": "Os atributos 'data_inicio' e 'data_fim' são obrigatórios"}), 400

    data_inicio = pd.to_datetime(inicio, format='%Y-%m-%d', errors='coerce')
    data_fim = pd.to_datetime(fim, format='%Y-%m-%d', errors='coerce')

    if pd.isna(data_inicio) or pd.isna(data_fim):
        return jsonify({"error": "Formato de data inválido. Use o formato 'YYYY-MM-DD'"}), 400

    if data_inicio > data_fim:
        return jsonify({"error": "A data de início deve ser anterior à data de fim"}), 400

    prontuarios = getatends(data_inicio, data_fim, operadoras)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pacientes"

    headers = ['PACIENTE', 'PRONTUARIO', 'ATENDIMENTO', 'DATA_INICIO', 'OPERADORA', 'STATUS', 'DATA_ALTA']
    ws.append(headers)

    for prontuario in prontuarios:
        ws.append([
            prontuario['PACIENTE'],
            prontuario['PRONTUARIO'],
            prontuario['ATENDIMENTO'],
            prontuario['DATA_INICIO'],
            prontuario['OPERADORA'],
            prontuario['STATUS'],
            prontuario['DATA_ALTA'],
        ])

    for cell in ws['B'] + ws['C']:
        cell.alignment = Alignment(horizontal='center')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=7):
        for cell in row:
            cell.number_format = 'DD/MM/YYYY HH:MM'
            cell.alignment = Alignment(horizontal='left')

    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = (max_length + 5)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Remover erros de validação
    ws.data_validations = []

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='dados.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )